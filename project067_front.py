import cv2
import pytesseract
import pandas as pd
from PIL import Image
import re
import PyPDF4
from openpyxl import load_workbook, Workbook

import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import os


subject_imgs = []
subject_index = 0
def browse_image():
    filetypes = (("Image files", "*.jpg;*.jpeg;*.png"), ("All files", "*.*"), ("PDF files", "*.pdf"))
    selected_files = filedialog.askopenfilenames(title="Select Image/PDF File", filetypes=filetypes)
    if selected_files:
        for file in selected_files:
            # Validate the image file
            if os.path.isfile(file):
                # Assign the image file to a backend variable
                global subject_imgs
                subject_imgs.append(file)
                print("Selected File:", file)
            else:
                print("Invalid image file.")


subject_excel = 'data.xlsx'


def execute_function_0():
    label_text6.config(text="Processing...")
    try:
        for subject_img in subject_imgs:
            InternalRef, County, Township, S, T, R, Lat, Long, Date, QSec = extract_table_data()  
            InternalRef_ = InternalRef
            County_ = County
            Township_ = Township
            S_ = S
            T_ = T
            R_ = R
            Latitude_ = Latitude
            Longitude_ = Longitude
            Date_ = Date
            QSec_ = QSec

            try:
            # Load the existing Excel file
                workbook = load_workbook(subject_excel)
                sheet = workbook[sheet_name]
                next_row = sheet.max_row + 1
            except FileNotFoundError:
                # If the file doesn't exist, create a new workbook
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = sheet_name
                next_row = 1

                column_headers = ['QuarterSection', 'Section(S)', 'Township(T)', 'Range(R)', 'County', 'Township', 'Date', 'Internal Ref.No.' , 'Latitude', 'Longitude']

                # Write the column headers to the first row
                for col_num, header in enumerate(column_headers, start=1):
                    sheet.cell(row=next_row, column=col_num, value=header)
                next_row += 1

            # Fill the data into the specific columns
            data_row = [ QSec_, S_, T_, R_, County_, Township_,  Date_, InternalRef_, Latitude_, Longitude_]
            for col_num, value in enumerate(data_row, start=1):
                sheet.cell(row=next_row, column=col_num, value=value)

            # Save the Excel file
            workbook.save(subject_excel)
            print('Excel file created successfully.')

            # Display "Finished" message
            label_text6.config(text="Finished")
            
    
        # Display a success message box
        messagebox.showinfo("Success", "Excel File has been Updated!")

    except Exception as e:
        # Display the error message in a messagebox
        messagebox.showerror("Error", str(e))

def execute_function():
    try:
        # Call the function and get the results
        InternalRef, County, Township, S, T, R, Lat, Long, Date, QSec = extract_table_data()  
        
        # Enable the Entry widgets for editing
        entry_value1.config(state=tk.NORMAL)
        entry_value1.delete(0, tk.END)
        entry_value1.insert(0, str(InternalRef))
        
        entry_value2.config(state=tk.NORMAL)
        entry_value2.delete(0, tk.END)
        entry_value2.insert(0, str(County))
        
        entry_value3.config(state=tk.NORMAL)
        entry_value3.delete(0, tk.END)
        entry_value3.insert(0, str(Township))

        entry_value4.config(state=tk.NORMAL)
        entry_value4.delete(0, tk.END)
        entry_value4.insert(0, str(S))

        entry_value5.config(state=tk.NORMAL)
        entry_value5.delete(0, tk.END)
        entry_value5.insert(0, str(T))

        entry_value6.config(state=tk.NORMAL)
        entry_value6.delete(0, tk.END)
        entry_value6.insert(0, str(R))

        entry_value7.config(state=tk.NORMAL)
        entry_value7.delete(0, tk.END)
        entry_value7.insert(0, str(Lat))

        entry_value8.config(state=tk.NORMAL)
        entry_value8.delete(0, tk.END)
        entry_value8.insert(0, str(Long))

        entry_value9.config(state=tk.NORMAL)
        entry_value9.delete(0, tk.END)
        entry_value9.insert(0, str(Date))

        entry_value10.config(state=tk.NORMAL)
        entry_value10.delete(0, tk.END)
        entry_value10.insert(0, str(QSec))
    except Exception as e:
        # Display the error message in a messagebox
        messagebox.showerror("Error", str(e))


InternalRef_ = None
County_ = None
Township_ = None
S_ = None
T_ = None
R_ = None
Latitude_ = None
Longitude_ = None
Date_ = None
QS_ = None

def update_values():
    # Get the new values from the Entry widgets
    new_value1 = entry_value1.get()
    new_value2 = entry_value2.get()
    new_value3 = entry_value3.get()
    new_value4 = entry_value4.get()
    new_value5 = entry_value5.get()
    new_value6 = entry_value6.get()
    new_value7 = entry_value7.get()
    new_value8 = entry_value8.get()
    new_value9 = entry_value9.get()
    new_value10 = entry_value10.get()

    global InternalRef_, County_, Township_, S_, T_, R_, Latitude_, Longitude_, Date_, QSec_
    InternalRef_= new_value1
    County_ = new_value2
    Township_ = new_value3
    S_ = new_value4
    T_ = new_value5
    R_ = new_value6
    Latitude_ = new_value7
    Longitude_ = new_value8
    Date_ = new_value9
    QSec_ = new_value10

    fill_excel(subject_excel, sheet_name)

    # Display a success message box
    messagebox.showinfo("Success", "Excel File has been Updated!")

    
def fill_excel(excel_path, sheet_name):
    # Load the existing Excel file
    try:
        global InternalRef_, County_, Township_, S_, T_, R_, Latitude_, Longitude_, Date_, QS_
        
        try:
        # Load the existing Excel file
            workbook = load_workbook(excel_path)
            sheet = workbook[sheet_name]
            next_row = sheet.max_row + 1
        except FileNotFoundError:
            # If the file doesn't exist, create a new workbook
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = sheet_name
            next_row = 1

            column_headers = ['QuarterSection', 'Section(S)', 'Township(T)', 'Range(R)', 'County', 'Township', 'Date', 'Internal Ref.No.' , 'Latitude', 'Longitude']

            # Write the column headers to the first row
            for col_num, header in enumerate(column_headers, start=1):
                sheet.cell(row=next_row, column=col_num, value=header)
            next_row += 1

        # Fill the data into the specific columns
        data_row = [ QS_, S_, T_, R_, County_, Township_,  Date_, InternalRef_, Latitude_, Longitude_]
        for col_num, value in enumerate(data_row, start=1):
            sheet.cell(row=next_row, column=col_num, value=value)

        # Save the Excel file
        workbook.save(excel_path)
        print('Excel file created successfully.')

        

    except Exception as e:
        # Display the error message in a messagebox
        messagebox.showerror("Error", str(e))

def is_image(file_path):
    # Get the file extension
    _, file_extension = os.path.splitext(file_path)

    # Check if the file extension corresponds to an image format
    image_extensions = ['.jpg', '.jpeg', '.png']
    return file_extension.lower() in image_extensions

def is_pdf(file_path):
    # Get the file extension
    _, file_extension = os.path.splitext(file_path)

    # Check if the file extension corresponds to a PDF format
    return file_extension.lower() == '.pdf'


def extract_table_data():
    global subject_index
    global subject_imgs
    try:
        if subject_index < len(subject_imgs):
            subject_img = subject_imgs[subject_index]

            if is_image(subject_img):

                image = Image.open(subject_img)
                ocr_text = pytesseract.image_to_string(image)
                print(ocr_text)
                
            elif is_pdf(subject_img):
                ocr_text = ''
                with open(subject_img, "rb") as file:
                    reader = PyPDF4.PdfFileReader(file)

                    for page_num in range(len(reader.pages)):
                        page = reader.pages[page_num]
                        ocr_text += page.extractText()
                

            else:
                subject_index += 1
                print("Unsupported File")
                ocr_text = None

            if ocr_text:
                pattern = r'\b(\w{2}-\w+-\d+-\d+)\b'
                pattern0 = r'\b(LX\s*-\S{1}[-—]\d+)\b'
                            
                LX_match = re.findall(pattern, ocr_text, re.I)
                for match in LX_match:
                    full_string = match.replace(" ", "").replace("\n", "")
                    print("Internal Ref.No:", full_string)
                    break  # Stop searching after the first match
                else:
                    LX_match1 = re.findall(pattern0, ocr_text)
                    for match in LX_match1:
                        full_string = match.replace(" ", "").replace("\n", "")
                        print("Internal Ref.No:", full_string)
                        break
                    else:   
                        full_string = 'Not Found'
                        print("Internal Ref.No.:", full_string)


                pattern1 = r'(\S+)\s+(?:COUNTY|CO)\b'    #r'(\S+)\s+COUNTY'   #r'[A-Z\s]+COUNTY'
                pattern11 = r'(?:Cook|Grundy|Kankakee|Lake|Cane|Mchenry)'
                county_match = re.findall(pattern1, ocr_text)
                if county_match:
                    last_match = county_match[-1]
                    county_string = last_match.replace(" ", "").replace("\n", "")
                    print("COUNTY:", county_string)
                else:
                    county1_match = re.findall(pattern11, ocr_text, re.I)
                    if county1_match:
                        last_match = county1_match[1]
                        county_string = last_match.replace(" ", "").replace("\n", "")
                        print("COUNTY:", county_string)
                    else:
                        county_string = 'Not Found'
                        print("COUNTY:", county_string)

                pattern2 = r'(\S+)\s*(?:TWP|TOWNSHIP)\b' #r'(?!\d+\s+)(.+?)\s+(?:TWP|TOWNSHIP)\b'  #   #r'(\S+)\s+TOWNSHIP'
                twp_match = re.findall(pattern2, ocr_text, re.I)
                if twp_match:
                    last_match = twp_match[-1]
                    twp_string = last_match.replace(" ", "").replace("\n", "")
                    print("TOWNSHIP:", twp_string)
                     # Stop searching after the first match
                else:
                    twp_string = 'Not Found'
                    print("TOWNSHIP:", twp_string)

                pattern3 = r'SEC[.\s]*(\d{1,2})\b'   #r'SEC[.\s]*\d+'
                pattern31 = r'SECTION: S(\d{2})' 
                sec_match = re.search(pattern3, ocr_text, re.I)
                if sec_match:
                    sec_string = sec_match.group(1)
                    print("SECTION :", sec_string)
                else:
                    sec1_match = re.search(pattern31, ocr_text)
                    if sec1_match:
                        sec_string = sec1_match.group(1)
                        print("SECTION :", sec_string)
                    else:
                        sec_string = 'Not Found'
                        print("SECTION :", sec_string)


                pattern4 = r'\bR[.\s]*([0-9]{1,2})[E|\d]*\b'
                R_match = re.search(pattern4, ocr_text)
                if R_match:
                    R_string = R_match.group(1)
                    print("R :", R_string)
                else:
                    R_string = 'Not Found'
                    print("R :", R_string)


                pattern5 = r'LAT: (\d{2}(?:[\s.]?\d+)*)'
                lat_match = re.search(pattern5, ocr_text, re.I)
                if lat_match:
                    lat_string = lat_match.group(1)
                    lat_value = lat_string.replace(' ', '.')
                    if (len(lat_value)<6):
                        lat_value = ''
                    print("LAT :", lat_value)
                else:
                    pattern51 = r'(\d{2}\.\d+)'   #r'(\d+(?:[\s.]?\d+)*,)'
                    lat_match = re.search(pattern51, ocr_text, re.I)
                    if lat_match:
                        lat_string = lat_match.group(1)
                        lat_value = lat_string.replace(' ', '.')
                        if (len(lat_value)<6):
                            lat_value = ''
                        print("LAT :", lat_value)
                    else:
                        lat_value = 'Not Found'
                        print("LAT :", lat_value)



                pattern6 = r'LONG: (-\d{2}(?:[\s.]?\d+)*)'  #r'LON(G)?: (-\d+(?:[\s.]?\d+)*)'
                long_match = re.search(pattern6, ocr_text, re.I)
                if long_match:
                    long_string = long_match.group(1)
                    long_value = long_string.replace(' ', '.')
                    if (len(long_value)<6):
                        long_value = ''
                    print("LONG :", long_value)

                else:
                    pattern61 = r'(-\d{2}\.\d+)'  #r'(-\d+(?:[\s.]?\d+)*)'  #r'LAT:\s*(-?\d+(?:[\s.]?\d+)*)'
                    long_match = re.search(pattern61, ocr_text, re.I)
                    if long_match:
                        long_string = long_match.group(1)
                        long_value = long_string.replace(' ', '.')
                        if (len(long_value)<6):
                            long_value = ''
                        print("LONG :", long_value)
                    else:
                        long_value = 'Not Found'
                        print("LONG :", long_value)


                pattern7 = r'(?:0[1-9]|1\d|2[0-9])[-/—](?:0[1-9]|1[0-2])[-/—]\d+' #r'(0[1-9]|1[0-2])[/\-](0[1-9]|[12][0-9]|3[01])[/\-]\d{4}'    #r'\d{2}[-/]\d{2}[-/]\d+'   #r'\d{2}/\d+/\d+'
                match = re.findall(pattern7, ocr_text)
                for date_match in match:
                    date = date_match[1]
                    date = date_match.replace(" ", "").replace("\n", "")
                    print("Date :", date)
                    break  # Stop searching after the first match
                else:
                    date = 'Not Found'
                    print("Date :", date)

                pattern8 = r'\b(?:\S{1,2}) (?:1/4|1/2)'   #r'\b(?:N(?:[EW])?|S(?:[EW])?|E|W)\s1[1/4|1/2]\b' 
                qs_match = re.search(pattern8, ocr_text)
                if qs_match:
                    qs_string = qs_match.group()
                    print("Quarter Section :", qs_string)
                else:
                    qs_string = 'Not Found'
                    print("Quarter Section :", qs_string)


                pattern9 = r'\b(?:TOWNSHIP|T|r)?\s*(\d{2})\s*N\b'
                twp_match = re.search(pattern9, ocr_text)
                if twp_match:
                    twp_value = twp_match.group(1)
                    #twp_string = last_match.replace(" ", "").replace("\n", "")
                    print("TOWNSHIP:", twp_value)
                else:
                    twp_value = 'Not Found'
                    print("TOWNSHIP:", twp_value)

                global InternalRef, County, Township, S, T, R, Latitude, Longitude, Date, QSec
                InternalRef = full_string
                County = county_string
                Township = twp_string
                S = sec_string
                T = twp_value
                R = R_string
                Latitude = lat_value
                Longitude = long_value
                Date = date
                QSec = qs_string

                subject_index += 1

                return full_string,county_string, twp_string, sec_string, twp_value, R_string, lat_value, long_value, date, qs_string
        elif subject_index == (len(subject_imgs)-1):
            print("No more images to Extract")
            subject_imgs.clear()

        else:

            print("No more images to Extract")

    except Exception as e:
        # Display the error message in a messagebox
        messagebox.showerror("Error", str(e))

# Create the Tkinter application window
window = tk.Tk()
window.title("SeleKta")
window.geometry("600x700")

subject_excel = 'data.xlsx'
sheet_name = 'Sheet1'

# Create a custom style for the buttons
button_style = {"background": "#87CEEB", "foreground": "white", "activebackground": "#45a049", "activeforeground": "white"}

# Create the Image File button
image_button = tk.Button(window, text="Select Image/PDF", command=browse_image, **button_style)
image_button.pack(pady=10)



extract_button = tk.Button(window, text="Extract", command=lambda: execute_function(), **button_style)
extract_button.pack(pady=10)

extract_button = tk.Button(window, text="Extract All", command=lambda: execute_function_0(), **button_style)
extract_button.pack(pady=10)

label_text6 = tk.Label(window, text="")
label_text6.pack()

label_text = tk.Label(window, text="Edit The Values Below If Incorrect")
label_text.pack()


label = tk.Label(window, text="InternalRefNo:")
label.pack()
# Create Entry widgets for editing the
entry_value1 = tk.Entry(window, state=tk.DISABLED)
entry_value1.pack()

label = tk.Label(window, text="County:")
label.pack()

entry_value2 = tk.Entry(window, state=tk.DISABLED)
entry_value2.pack()

label = tk.Label(window, text="Township")
label.pack()

entry_value3 = tk.Entry(window, state=tk.DISABLED)
entry_value3.pack()

label = tk.Label(window, text="Section Township Range")
label.pack()

entry_value4 = tk.Entry(window, state=tk.DISABLED)
entry_value4.pack()

entry_value5 = tk.Entry(window, state=tk.DISABLED)
entry_value5.pack()

entry_value6 = tk.Entry(window, state=tk.DISABLED)
entry_value6.pack()

label = tk.Label(window, text="Latitude $ Longitude")
label.pack()

entry_value7 = tk.Entry(window, state=tk.DISABLED)
entry_value7.pack()

entry_value8 = tk.Entry(window, state=tk.DISABLED)
entry_value8.pack()

label = tk.Label(window, text="Date")
label.pack()

entry_value9 = tk.Entry(window, state=tk.DISABLED)
entry_value9.pack()

label = tk.Label(window, text="Quarter Section")
label.pack()

entry_value10 = tk.Entry(window, state=tk.DISABLED)
entry_value10.pack()

label_text2 = tk.Label(window, text="Update Values \'ONCE\'")
label_text2.pack()

# Create a button to update the values
update_button = tk.Button(window, text="Update Values", command=update_values, **button_style)
update_button.pack()




# Run the Tkinter event loop
window.mainloop()